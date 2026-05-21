from openpyxl import Workbook
from datetime import datetime


def crt_inventory_subtotal_xlsx(inventory_subtotal: list, warehouse: list) -> str:
    """
    產生庫存樞紐excel
    :param inventory_subtotal: 庫存樞紐資料
    :param warehouse: 倉別列表
    """
    workbook = Workbook()
    sheet1 = workbook.active
    sheet1.title = "庫存樞紐"
    
    # 寫入表頭
    headers = ['料號', '品名'] + warehouse + ['庫存總計', 'GS庫存數量', '庫存差異數']
    for col_idx, header in enumerate(headers, start=1):
        sheet1.cell(row=1, column=col_idx, value=header)
    
    # 寫入資料
    for row_idx, item in enumerate(inventory_subtotal, start=2):
        # 料號
        sheet1.cell(row=row_idx, column=1, value=item.get('s_item', ''))
        
        # 品名
        sheet1.cell(row=row_idx, column=2, value=item.get('s_item_name', ''))
        
        # 各倉別數量
        for col_idx, wh in enumerate(warehouse, start=3):
            value = item.get(wh)
            # 如果是 0 或 None，顯示空白
            if value is None or value == 0 or value == 0.0:
                display_value = ''
            else:
                display_value = value
            sheet1.cell(row=row_idx, column=col_idx, value=display_value)
        
        # 庫存總計
        total_value = item.get('total_inventory', 0)
        # 總計轉為整數
        if total_value is not None:
            total_value = round(float(total_value))
        else:
            total_value = 0
        sheet1.cell(row=row_idx, column=len(warehouse) + 3, value=total_value)

        # GS庫存數量
        gs_value = item.get('total_gs_num', 0)
        # GS庫存數量轉為整數
        if gs_value is not None:
            gs_value = round(float(gs_value))
        else:
            gs_value = 0
        sheet1.cell(row=row_idx, column=len(warehouse) + 4, value=gs_value)

        # 庫存差異數
        diff_value = item.get('inventory_difference', 0)
        # 差異數轉為整數
        if diff_value is not None:
            diff_value = round(float(diff_value))
        else:
            diff_value = 0
        sheet1.cell(row=row_idx, column=len(warehouse) + 5, value=diff_value)   
    # 儲存檔案
    file_name: str = f"庫存樞紐_{str(datetime.now().strftime('%Y%m%d_%H%M%S'))}.xlsx"
    workbook.save(file_name)
    return file_name